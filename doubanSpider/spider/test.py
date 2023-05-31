#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time    : 2023/5/25 0:58
# @Author  : Smalltown
# @FileName: test.py
# @Software: PyCharm
import xlwt
import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
def saveurl_drama(datalist,savepath,num):
    print("save...")
    book = xlwt.Workbook(encoding='utf-8', style_compression=0)  # 创建workbook对象
    sheet = book.add_sheet('豆瓣电影剧情片url',cell_overwrite_ok=True)  # 创建工作表
    col = ("电影详情链接")
    sheet.write(0,0,col) #列名
    for i in range(0,num):
        print("第%d条"%(i+1))
        data = datalist[i]
        sheet.write(i+1,0,data)  #数据


    book.save(savepath)  # 保存数据表


datalist = ['https://www.douban.com/doubanapp/dispatch?uri=/movie/27066152',
'https://www.douban.com/doubanapp/dispatch?uri=/movie/4811807',
'https://www.douban.com/doubanapp/dispatch?uri=/movie/2161696',
'https://www.douban.com/doubanapp/dispatch?uri=/movie/26924143'

]

# saveurl_drama(datalist,savepath = ".\\豆瓣电影test.xls",num=4)



# data =[{'rating': ['9.7', '50'], 'rank': 1, 'cover_url': 'https://img2.doubanio.com/view/photo/s_ratio_poster/public/p480747492.jpg', 'is_playable': True, 'id': '1292052', 'types': ['犯罪', '剧情'], 'regions': ['美国'], 'title': '肖申克的救赎', 'url': 'https://movie.douban.com/subject/1292052/', 'release_date': '1994-09-10', 'actor_count': 25, 'vote_count': 2857558, 'score': '9.7', 'actors': ['蒂姆·罗宾斯', '摩根·弗里曼', '鲍勃·冈顿', '威廉姆·赛德勒', '克兰西·布朗', '吉尔·贝罗斯', '马克·罗斯顿', '詹姆斯·惠特摩', '杰弗里·德曼', '拉里·布兰登伯格', '尼尔·吉恩托利', '布赖恩·利比', '大卫·普罗瓦尔', '约瑟夫·劳格诺', '祖德·塞克利拉', '保罗·麦克兰尼', '芮妮·布莱恩', '阿方索·弗里曼', 'V·J·福斯特', '弗兰克·梅德拉诺', '马克·迈尔斯', '尼尔·萨默斯', '耐德·巴拉米', '布赖恩·戴拉特', '唐·麦克马纳斯'], 'is_watched': False}, {'rating': ['9.6', '50'], 'rank': 2, 'cover_url': 'https://img1.doubanio.com/view/photo/s_ratio_poster/public/p2561716440.jpg', 'is_playable': True, 'id': '1291546', 'types': ['剧情', '爱情', '同性'], 'regions': ['中国大陆', '中国香港'], 'title': '霸王别姬', 'url': 'https://movie.douban.com/subject/1291546/', 'release_date': '1993-07-26', 'actor_count': 27, 'vote_count': 2112726, 'score': '9.6', 'actors': ['张国荣', '张丰毅', '巩俐', '葛优', '英达', '蒋雯丽', '吴大维', '吕齐', '雷汉', '尹治', '马明威', '费振翔', '智一桐', '李春', '赵海龙', '李丹', '童弟', '沈慧芬', '黄斐', '徐杰', '黄磊', '冯远征', '杨立新', '方征', '周璞', '隋永清', '宋小川'], 'is_watched': False}, {'rating': ['9.6', '50'], 'rank': 3, 'cover_url': 'https://img2.doubanio.com/view/photo/s_ratio_poster/public/p2578474613.jpg', 'is_playable': True, 'id': '1292063', 'types': ['剧情', '喜剧', '爱情', '战争'], 'regions': ['意大利'], 'title': '美丽人生', 'url': 'https://movie.douban.com/subject/1292063/', 'release_date': '2020-01-03', 'actor_count': 29, 'vote_count': 1310480, 'score': '9.6', 'actors': ['罗伯托·贝尼尼', '尼可莱塔·布拉斯基', '乔治·坎塔里尼', '朱斯蒂诺·杜拉诺', '赛尔乔·比尼·布斯特里克', '玛丽萨·帕雷德斯', '霍斯特·布赫霍尔茨', '利迪娅·阿方西', '朱利亚娜·洛约迪切', '亚美利哥·丰塔尼', '彼得·德·席尔瓦', '弗朗西斯·古佐', '拉法埃拉·莱博罗尼', '克劳迪奥·阿方西', '吉尔·巴罗尼', '马西莫·比安奇', '恩尼奥·孔萨尔维', '吉安卡尔洛·科森蒂诺', '阿伦·克雷格', '汉尼斯·赫尔曼', '弗兰科·梅斯科利尼', '安东尼奥·普雷斯特', '吉娜·诺维勒', '理查德·塞梅尔', '安德烈提多娜', '迪尔克·范登贝格', '奥梅罗·安东努蒂', '沈晓谦', '张欣'], 'is_watched': False}, {'rating': ['9.6', '50'], 'rank': 4, 'cover_url': 'https://img2.doubanio.com/view/photo/s_ratio_poster/public/p492406163.jpg', 'is_playable': True, 'id': '1295124', 'types': ['剧情', '历史', '战争'], 'regions': ['美国'], 'title': '辛德勒的名单', 'url': 'https://movie.douban.com/subject/1295124/', 'release_date': '1993-11-30', 'actor_count': 49, 'vote_count': 1091983, 'score': '9.6', 'actors': ['连姆·尼森', '本·金斯利', '拉尔夫·费因斯', '卡罗琳·古多尔', '乔纳森·萨加尔', '艾伯丝·戴维兹', '马尔戈萨·格贝尔', '马克·伊瓦涅', '碧翠斯·马科拉', '安德烈·瑟韦林', '弗里德里希·冯·图恩', '克齐斯茨托夫·拉夫特', '诺伯特·魏塞尔', '维斯瓦夫·科马萨', '布拉德·雅各布维茨', 'Maciej Orlos', '皮奥特·赛尔沃斯', 'Tadeusz Huk', '马丁·塞梅洛格', '托马斯·德德克', '奥拉夫·卢巴申科', '马瑞安·格林卡', '约亨·尼克尔', '阿格涅兹卡·克鲁科沃娜', '阿格尼兹卡·旺格', '托马斯·莫里斯', '佐久间玲', '吴俊全', '约阿希姆·保罗·阿斯波克', '彭河', '戈兹·奥托', '玛雅·奥丝塔泽斯卡', 'Maciej Kozlowski', '艾尔文·莱德', 'Eugeniusz Priwieziencew', 'Marta Bizon', '埃兹拉·达甘', '吉恩·莱赫纳', 'Razia Israeli', '拉米·希尔伯格', '布兰科·拉斯蒂格', '路德格·皮斯特', '埃琳娜·勒文松', '胡契克·卡勒塔', '塔德乌什·布拉德茨基', '亨里克·比斯塔', '帕维·德朗柯', '耶日·诺瓦克', '安娜·穆查'], 'is_watched': False}, {'rating': ['9.6', '50'], 'rank': 5, 'cover_url': 'https://img1.doubanio.com/view/photo/s_ratio_poster/public/p1505392928.jpg', 'is_playable': True, 'id': '1296141', 'types': ['剧情', '犯罪', '悬疑'], 'regions': ['美国'], 'title': '控方证人', 'url': 'https://movie.douban.com/subject/1296141/', 'release_date': '1957-12-17', 'actor_count': 51, 'vote_count': 542743, 'score': '9.6', 'actors': ['泰隆·鲍华', '玛琳·黛德丽', '查尔斯·劳顿', '埃尔莎·兰彻斯特', '约翰·威廉姆斯', '亨利·丹尼尔', '伊安·沃尔夫', '托林·撒切尔', '诺玛·威登', '尤娜·奥康纳', '茹塔·李', '贝丝·弗劳尔斯', '比尔·厄尔文', 'J·帕特·奥马利', '本·怀特', 'Paul Kruger', 'Jack Raine', 'Paul Power', '乔治·佩林', '威廉·H·奥布莱恩', '奥托拉内史密斯', 'Frank McClure', 'Colin Kenny', 'Jeanne Lafayette', 'Wilbur Mack', 'Fred Rapport', '利奥达·理查德斯', 'Glen Walters', 'Arthur Tovey', '伯特史蒂文斯', 'Cap Somers', 'Lucille Sewall', '斯考特西顿', 'Norbert Schiller', '杰弗里·塞尔', 'John Roy', 'Al Roberts', 'Art Howard', 'Stuart Hall', 'Francis Compton', 'Philip Tonge', '帕特里克·艾亨', '富兰克林·法纳姆', '玛乔丽·伊顿', '史蒂夫·卡鲁瑟斯', 'George Calliga', '乔治布鲁格曼', '丹尼·鲍沙其', 'Brandon Beach', '埃迪·贝克', '沃尔特·培根'], 'is_watched': False}, {'rating': ['9.5', '50'], 'rank': 6, 'cover_url': 'https://img9.doubanio.com/view/photo/s_ratio_poster/public/p2889314814.jpg', 'is_playable': True, 'id': '1292722', 'types': ['剧情', '爱情', '灾难'], 'regions': ['美国', '墨西哥'], 'title': '泰坦尼克号', 'url': 'https://movie.douban.com/subject/1292722/', 'release_date': '1998-04-03', 'actor_count': 63, 'vote_count': 2158837, 'score': '9.5', 'actors': ['莱昂纳多·迪卡普里奥', '凯特·温丝莱特', '比利·赞恩', '凯西·贝茨', '弗兰西丝·费舍', '格劳瑞亚·斯图尔特', '比尔·帕克斯顿', '伯纳德·希尔', '大卫·沃纳', '维克多·加博', '乔纳森·海德', '苏茜·爱米斯', '刘易斯·阿伯内西', '尼古拉斯·卡斯柯恩', '阿那托利·萨加洛维奇', '丹尼·努齐', '杰森·贝瑞', '伊万·斯图尔特', '艾恩·格拉法德', '乔纳森·菲利普斯', '马克·林赛·查普曼', '理查德·格拉翰', '保罗·布赖特威尔', '艾瑞克·布里登', '夏洛特·查顿', '博纳德·福克斯', '迈克尔·英塞恩', '法妮·布雷特', '马丁·贾维斯', '罗莎琳·艾尔斯', '罗切尔·罗斯', '乔纳森·伊万斯-琼斯', '西蒙·克雷恩', '爱德华德·弗莱彻', '斯科特·安德森', '马丁·伊斯特', '克雷格·凯利', '格雷戈里·库克', '利亚姆·图伊', '詹姆斯·兰开斯特', '艾尔莎·瑞雯', '卢·帕尔特', '泰瑞·佛瑞斯塔', '凯文·德·拉·诺伊', '詹姆斯·卡梅隆', '竹内结子', '詹姆斯·霍纳', '姜广涛', '罗恩·多纳基', '佐佐木澄江', '史蒂文·奎里', '詹妮特·戈德斯坦恩', '林凡 ', 'Linda Kerns', '马丁·哈伯', '马丁·莱恩', '德里克·李', '保罗·赫伯特', '艾利克斯·欧文斯-萨尔诺', 'Julene Renee', '刘之玲', '曲敬国', '季冠霖'], 'is_watched': False}, {'rating': ['9.5', '50'], 'rank': 7, 'cover_url': 'https://img2.doubanio.com/view/photo/s_ratio_poster/public/p2372307693.jpg', 'is_playable': True, 'id': '1292720', 'types': ['剧情', '爱情'], 'regions': ['美国'], 'title': '阿甘正传', 'url': 'https://movie.douban.com/subject/1292720/', 'release_date': '1994-06-23', 'actor_count': 51, 'vote_count': 2133125, 'score': '9.5', 'actors': ['汤姆·汉克斯', '罗宾·怀特', '加里·西尼斯', '麦凯尔泰·威廉逊', '莎莉·菲尔德', '海利·乔·奥斯蒙', '迈克尔·康纳·亨弗里斯', '哈罗德·G·赫瑟姆', '山姆·安德森', '伊俄涅·M·特雷奇', '彼得·道博森', '希博汗·法隆', '伊丽莎白·汉克斯', '汉娜·豪尔', '克里斯托弗·琼斯', '罗布·兰德里', '杰森·麦克奎尔', '桑尼·施罗耶', '艾德·戴维斯', '丹尼尔C.斯瑞派克', '大卫·布里斯宾', '德博拉·麦克蒂尔', '艾尔·哈林顿', '阿非莫·奥米拉', '约翰·沃德斯塔德', '迈克尔·伯吉斯', '埃里克·安德伍德', '拜伦·明斯', '斯蒂芬·布吉格沃特', '约翰·威廉·高尔特', '希拉里·沙普兰', '伊莎贝尔·罗斯', '理查德·达历山德罗', '迪克·史迪威', '迈克尔-杰斯', '杰弗里·布莱克', '瓦妮莎·罗斯', '迪克·卡维特', '马拉·苏查雷特扎', '乔·阿拉斯奇', 'W·本森·泰瑞', '约翰·列侬', '埃尔维斯·普雷斯利', '罗纳德·里根', '鲍勃·霍普', '约翰·肯尼迪', '理查德·尼克松', '吴俊全', '蒂莫西·麦克尼尔', '杰拉尔德·福特', '布雷特·赖斯'], 'is_watched': False}, {'rating': ['9.6', '50'], 'rank': 8, 'cover_url': 'https://img9.doubanio.com/view/photo/s_ratio_poster/public/p2528965424.jpg', 'is_playable': True, 'id': '1461403', 'types': ['剧情', '历史'], 'regions': ['中国大陆'], 'title': '茶馆', 'url': 'https://movie.douban.com/subject/1461403/', 'release_date': '1982', 'actor_count': 36, 'vote_count': 139436, 'score': '9.6', 'actors': ['于是之', '郑榕', '蓝天野', '英若诚', '黄宗洛', '童超', '金昭', '林连昆', '牛星丽', '谭宗尧', '童弟', '吴淑昆', '胡宗温', '程中', '李大千', '李源', '李翔', '张瞳', '任宝贤', '尚丽娟', '米铁增', '田春奎', '冯增祥', '雷飞', '林东升', '孟瑾', '孙峻峰', '平原', '曹世骧', '王大年', '王淑华', '张华', '丁海涛', '朱旭', '孙敬修', '董行佶'], 'is_watched': False}, {'rating': ['9.4', '50'], 'rank': 9, 'cover_url': 'https://img2.doubanio.com/view/photo/s_ratio_poster/public/p511118051.jpg', 'is_playable': True, 'id': '1295644', 'types': ['剧情', '动作', '犯罪'], 'regions': ['法国', '美国'], 'title': '这个杀手不太冷', 'url': 'https://movie.douban.com/subject/1295644/', 'release_date': '1994-09-14', 'actor_count': 22, 'vote_count': 2270266, 'score': '9.4', 'actors': ['让·雷诺', '娜塔莉·波特曼', '加里·奥德曼', '丹尼·爱罗', '彼得·阿佩尔', '迈克尔·巴达鲁科', '艾伦·格里尼', '伊丽莎白·瑞根', '卡尔·马图斯维奇', '弗兰克·赛格', '麦温', '乔治·马丁', '罗伯特·拉萨多', '亚当·布斯奇', '马里奥·托迪斯科', '萨米·纳塞利', '让·雨果·安格拉德', '埃莱娜·卡多纳', '沈晓谦', '大塚明夫', 'Keith A. Glascoe', 'Arsène Jiroyan'], 'is_watched': False}, {'rating': ['9.4', '50'], 'rank': 10, 'cover_url': 'https://img1.doubanio.com/view/photo/s_ratio_poster/public/p2557573348.jpg', 'is_playable': False, 'id': '1291561', 'types': ['剧情', '动画', '奇幻'], 'regions': ['日本'], 'title': '千与千寻', 'url': 'https://movie.douban.com/subject/1291561/', 'release_date': '2019-06-21', 'actor_count': 26, 'vote_count': 2213946, 'score': '9.4', 'actors': ['柊瑠美', '入野自由', '夏木真理', '菅原文太', '中村彰男', '玉井夕海', '神木隆之介', '内藤刚志', '泽口靖子', '我修院达也', '大泉洋', '小林郁夫', '上条恒彦', '小野武彦', '田壮壮', '王琳', '安田显', '户次重幸', '胡立成', '山像香', '斋藤志郎', '脇田茂', '彭昱畅', '井柏然', '周冬雨', '毛毛头'], 'is_watched': False}, {'rating': ['9.4', '50'], 'rank': 11, 'cover_url': 'https://img9.doubanio.com/view/photo/s_ratio_poster/public/p513344864.jpg', 'is_playable': True, 'id': '3541415', 'types': ['剧情', '科幻', '悬疑', '冒险'], 'regions': ['美国', '英国'], 'title': '盗梦空间', 'url': 'https://movie.douban.com/subject/3541415/', 'release_date': '2010-09-01', 'actor_count': 30, 'vote_count': 2035043, 'score': '9.4', 'actors': ['莱昂纳多·迪卡普里奥', '约瑟夫·高登-莱维特', '艾利奥特·佩吉', '汤姆·哈迪', '渡边谦', '迪利普·劳', '基里安·墨菲', '汤姆·贝伦杰', '玛丽昂·歌迪亚', '皮特·波斯尔思韦特', '迈克尔·凯恩', '卢卡斯·哈斯', '李太力', '克莱尔·吉尔蕾', '马格努斯·诺兰', '泰勒·吉蕾', '乔纳森·吉尔', '水源士郎', '冈本玉二', '厄尔·卡梅伦', '瑞恩·海沃德', '米兰达·诺兰', '拉什·费加', '蒂姆·科勒赫', '妲露拉·莱莉', '迈克尔·加斯顿', '吉尔·马德雷尔', '玛格丽特·因索利亚', '布拉德·雅各布维茨', 'Natasha Beaumont'], 'is_watched': False}, {'rating': ['9.4', '50'], 'rank': 12, 'cover_url': 'https://img1.doubanio.com/view/photo/s_ratio_poster/public/p2614988097.jpg', 'is_playable': True, 'id': '1889243', 'types': ['剧情', '科幻', '冒险'], 'regions': ['美国', '英国', '加拿大'], 'title': '星际穿越', 'url': 'https://movie.douban.com/subject/1889243/', 'release_date': '2014-11-12', 'actor_count': 28, 'vote_count': 1803623, 'score': '9.4', 'actors': ['马修·麦康纳', '安妮·海瑟薇', '杰西卡·查斯坦', '麦肯吉·弗依', '卡西·阿弗莱克', '迈克尔·凯恩', '马特·达蒙', '提莫西·查拉梅', '艾伦·伯斯汀', '约翰·利思戈', '韦斯·本特利', '大卫·吉亚西', '比尔·欧文', '托弗·戈瑞斯', '科莱特·沃夫', '弗朗西斯·X.麦卡蒂', '安德鲁·博尔巴', '乔什·斯图沃特', '莱雅·卡里恩斯', '利亚姆·迪金森', '杰夫·赫普内尔', '伊莱耶斯·加贝尔', '布鲁克·史密斯', '大卫·奥伊罗', '威廉姆·德瓦内', '拉什·费加', '格里芬·弗雷泽', '弗洛拉·诺兰'], 'is_watched': False}, {'rating': ['9.4', '50'], 'rank': 13, 'cover_url': 'https://img2.doubanio.com/view/photo/s_ratio_poster/public/p479682972.jpg', 'is_playable': True, 'id': '1292064', 'types': ['剧情', '科幻'], 'regions': ['美国'], 'title': '楚门的世界', 'url': 'https://movie.douban.com/subject/1292064/', 'release_date': '1998-06-05', 'actor_count': 26, 'vote_count': 1676077, 'score': '9.4', 'actors': ['金·凯瑞', '劳拉·琳妮', '艾德·哈里斯', '诺亚·艾默里奇', '娜塔莎·麦克艾霍恩', '安东尼·科隆', '马西娅·德波尼斯', 'Adam Tomei', '约翰·普莱舍', '澳澜·琼斯', 'Joe Minjares', '特里·金瑞利', '乔尔·麦金农·米勒', '冈本玉二', 'Jake Eberle', '马克·麦考利', '贾德森·沃恩', '彼得·克劳斯', '保罗·吉亚玛提', '菲利普·贝克·霍尔', '梅丽莎·菲茨杰拉德', 'Blair Slater', '珍妮特·米勒', '霍兰德·泰勒', '布赖恩·戴拉特', '哈里·谢尔'], 'is_watched': False}, {'rating': ['9.4', '50'], 'rank': 14, 'cover_url': 'https://img1.doubanio.com/view/photo/s_ratio_poster/public/p2587099240.jpg', 'is_playable': False, 'id': '3011091', 'types': ['剧情'], 'regions': ['美国', '英国'], 'title': '忠犬八公的故事', 'url': 'https://movie.douban.com/subject/3011091/', 'release_date': '2009-06-13', 'actor_count': 22, 'vote_count': 1384326, 'score': '9.4', 'actors': ['理查·基尔', '莎拉·罗默尔', '琼·艾伦', '罗比·萨布莱特', '艾瑞克·阿瓦利', '田川洋行', '杰森·亚历山大', '罗伯特·卡普荣', '凯文·德科斯泰', '高岛彩', '奇科', '莱拉', 'Ellen Becker-Gray', 'Ben Skinner', 'Raymond Alongi', '达文尼亚·麦克法登', '贝茨·怀尔德', '亚美利哥·普雷休蒂', '盖尔·布格贾', '马丁·蒙塔纳', 'Michael Kelly', '福雷斯特'], 'is_watched': False}, {'rating': ['9.4', '50'], 'rank': 15, 'cover_url': 'https://img9.doubanio.com/view/photo/s_ratio_poster/public/p1363250216.jpg', 'is_playable': False, 'id': '5912992', 'types': ['剧情'], 'regions': ['韩国'], 'title': '熔炉', 'url': 'https://movie.douban.com/subject/5912992/', 'release_date': '2011-09-22', 'actor_count': 22, 'vote_count': 920796, 'score': '9.4', 'actors': ['孔刘', '郑有美', '金贤秀', '郑仁絮', '白承焕', '张光', '金民尚', '林贤成', '金周灵', '严孝燮', '全国焕', '崔镇浩', '金志映', '严智星', '许在浩', '张素妍', '金智怜', '张宥', '朴惠珍', '林贤成', '郑亨锡', '洪锡然'], 'is_watched': False}, {'rating': ['9.4', '50'], 'rank': 16, 'cover_url': 'https://img1.doubanio.com/view/photo/s_ratio_poster/public/p2184505167.jpg', 'is_playable': True, 'id': '1418019', 'types': ['剧情', '动画', '奇幻', '古装'], 'regions': ['中国大陆'], 'title': '大闹天宫', 'url': 'https://movie.douban.com/subject/1418019/', 'release_date': '1961', 'actor_count': 7, 'vote_count': 430101, 'score': '9.4', 'actors': ['邱岳峰', '富润生', '毕克', '尚华', '于鼎', '李梓', '刘广宁'], 'is_watched': False}, {'rating': ['9.4', '50'], 'rank': 17, 'cover_url': 'https://img2.doubanio.com/view/photo/s_ratio_poster/public/p2173577632.jpg', 'is_playable': True, 'id': '1293182', 'types': ['剧情'], 'regions': ['美国'], 'title': '十二怒汉', 'url': 'https://movie.douban.com/subject/1293182/', 'release_date': '1957-04-10', 'actor_count': 12, 'vote_count': 482177, 'score': '9.4', 'actors': ['亨利·方达', '马丁·鲍尔萨姆', '约翰·菲德勒', '李·科布', 'E.G.马绍尔', '杰克·克卢格曼', '爱德华·宾斯', '杰克·瓦尔登', '约瑟夫·史威尼', '埃德·贝格利', '乔治·沃斯科维奇', '罗伯特·韦伯'], 'is_watched': False}, {'rating': ['9.3', '50'], 'rank': 18, 'cover_url': 'https://img2.doubanio.com/view/photo/s_ratio_poster/public/p2564556863.jpg', 'is_playable': True, 'id': '1307914', 'types': ['剧情', '犯罪', '惊悚'], 'regions': ['中国香港'], 'title': '无间道', 'url': 'https://movie.douban.com/subject/1307914/', 'release_date': '2003-09-05', 'actor_count': 30, 'vote_count': 1343401, 'score': '9.3', 'actors': ['刘德华', '梁朝伟', '黄秋生', '曾志伟', '陈慧琳', '郑秀文', '陈冠希', '余文乐', '萧亚轩', '林家栋', '吴廷烨', '林迪安', '尹志强', '许金峰', '何华超', '利沙华', '区轩玮', '李天翔', '黄燕强', '姚文基', '余世腾', '苏伟南', '黎志伟', '梁皓楷', '张旭燊', '袁伟豪', '叶清', '洪智杰', '张艺', '杨容莲'], 'is_watched': False}, {'rating': ['9.5', '50'], 'rank': 19, 'cover_url': 'https://img1.doubanio.com/view/photo/s_ratio_poster/public/p2505048077.jpg', 'is_playable': True, 'id': '1307856', 'types': ['剧情'], 'regions': ['中国大陆', '中国香港'], 'title': '背靠背，脸对脸', 'url': 'https://movie.douban.com/subject/1307856/', 'release_date': '1994-09-10', 'actor_count': 18, 'vote_count': 128860, 'score': '9.5', 'actors': ['牛振华', '雷恪生', '李强', '句号', '王劲松', '戈治均', '佘南南', '张嘉益', '刘国祥', '徐学政', '林海海', '温谦', '马载军', '吴玮玲', '张艺', '韩炳杰', '杨亚洲', '王洪光'], 'is_watched': False}, {'rating': ['9.3', '50'], 'rank': 20, 'cover_url': 'https://img9.doubanio.com/view/photo/s_ratio_poster/public/p2574551676.jpg', 'is_playable': True, 'id': '1292001', 'types': ['剧情', '音乐'], 'regions': ['意大利'], 'title': '海上钢琴师', 'url': 'https://movie.douban.com/subject/1292001/', 'release_date': '2019-11-15', 'actor_count': 23, 'vote_count': 1657959, 'score': '9.3', 'actors': ['蒂姆·罗斯', '普路特·泰勒·文斯', '比尔·努恩', '克兰伦斯·威廉姆斯三世', '梅兰尼·蒂埃里', '皮特·沃恩', '尼尔·奥布赖恩', '阿尔贝托·巴斯克斯', '加布里埃莱·拉维亚', '科里·巴克', '西德尼·科尔', 'Luigi De Luca', '尼古拉·迪·平托', '费米·依鲁福祖', '伊斯顿·盖奇', '凯文·麦克纳利', '布莱恩·普林格', '沙拉·鲁宾', '希思科特·威廉姆斯', '阿妮妲·扎格利亚', '安吉洛·迪洛雷塔', '吉达·布塔', '曲敬国'], 'is_watched': False}]
# print(len(data))
def saveurl(url_list, save_path):
    data = {"URL": url_list}
    df = pd.DataFrame(data)
    df.to_excel(save_path, index=False)
    print("URLs saved to", save_path)



from openpyxl import load_workbook
from openpyxl import Workbook

def append_to_excel(data_list, save_path):
    try:
        workbook = load_workbook(save_path)  # 加载现有的 Excel 文件
        sheet = workbook.active  # 获取活动工作表
        last_row = sheet.max_row  # 获取最后一行的行号
        start_row = last_row + 1  # 下一行开始写入的行号

        for data in data_list:
            sheet.append([data])  # 将数据封装到列表中并追加到工作表的下一行

        workbook.save(save_path)  # 保存更改
        workbook.close()  # 关闭工作簿
        print("Data appended to", save_path)
    except FileNotFoundError:
        print("Excel file not found. Creating a new file...")
        workbook = Workbook()  # 创建一个新的工作簿
        sheet = workbook.active  # 获取活动工作表
        sheet.insert_cols(1)  # 在第一列之前插入新列
        sheet.cell(row=1, column=1, value="电影链接URL")  # 设置第一行第一列的单元格值
        for data in data_list:
            sheet.append([data])  # 将数据封装到列表中并追加到工作表的下一行
        workbook.save(save_path)  # 保存工作簿
        workbook.close()  # 关闭工作簿
        print("New Excel file created and data appended to", save_path)
    except Exception as e:
        print("An error occurred while appending data:", e)

# 调用示例：
savepath = "urlstest.xlsx"
url_list = ['https://movie.douban.com/subject/1292052/', 'https://movie.douban.com/subject/1291546/', 'https://movie.douban.com/subject/1292063/', 'https://movie.douban.com/subject/1295124/', 'https://movie.douban.com/subject/1296141/', 'https://movie.douban.com/subject/1292722/', 'https://movie.douban.com/subject/1292720/', 'https://movie.douban.com/subject/1461403/', 'https://movie.douban.com/subject/1295644/', 'https://movie.douban.com/subject/1291561/', 'https://movie.douban.com/subject/3541415/', 'https://movie.douban.com/subject/1889243/', 'https://movie.douban.com/subject/1292064/', 'https://movie.douban.com/subject/3011091/', 'https://movie.douban.com/subject/5912992/', 'https://movie.douban.com/subject/1418019/', 'https://movie.douban.com/subject/1293182/', 'https://movie.douban.com/subject/1307914/', 'https://movie.douban.com/subject/1307856/', 'https://movie.douban.com/subject/1292001/']
  # 假设这是你获取的链接列表
# append_to_excel(url_list, savepath)
saveurl(url_list,savepath)
# url_list2 =['https://movie.douban.com/subject/1293172/', 'https://movie.douban.com/subject/1292365/', 'https://movie.douban.com/subject/1291549/', 'https://movie.douban.com/subject/1291858/', 'https://movie.douban.com/subject/6786002/', 'https://movie.douban.com/subject/1291841/', 'https://movie.douban.com/subject/1291552/', 'https://movie.douban.com/subject/1300267/', 'https://movie.douban.com/subject/21937452/', 'https://movie.douban.com/subject/1296736/', 'https://movie.douban.com/subject/25958717/', 'https://movie.douban.com/subject/1299131/', 'https://movie.douban.com/subject/3793023/', 'https://movie.douban.com/subject/1849031/', 'https://movie.douban.com/subject/1291818/', 'https://movie.douban.com/subject/1851857/', 'https://movie.douban.com/subject/1291572/', 'https://movie.douban.com/subject/1294371/', 'https://movie.douban.com/subject/1291548/', 'https://movie.douban.com/subject/1422283/']
# append_to_excel(url_list2, savepath)