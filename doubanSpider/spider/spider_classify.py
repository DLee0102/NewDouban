#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time    : 2023/5/24 23:10
# @Author  : Smalltown
# @FileName: spider_classify.py
# @Software: PyCharm
import re
'''
    爬取url的模块文件
'''
import json
import random
import time
import urllib

import urllib.request, urllib.error     #制定URL，获取网页数据
# excel操作
from openpyxl import load_workbook
from openpyxl import Workbook
'''
100-90
https://movie.douban.com/j/chart/top_list?type=11&interval_id=100%3A90&action=&start=0&limit=20
https://movie.douban.com/j/chart/top_list?type=11&interval_id=100%3A90&action=&start=20&limit=20
……
90-80
https://movie.douban.com/j/chart/top_list?type=  11  &interval_id=  90%3A80  &action=&start=  0  &limit=20
https://movie.douban.com/j/chart/top_list?type=11&interval_id=90%3A80&action=&start=20&limit=20

剧情：11
喜剧：24
动作：5
爱情：13

'''
baseurl_left = r"https://movie.douban.com/j/chart/top_list?type="
baseurl_midd1 = r"&interval_id="

baseurl_midd3 = r"&action=&start="
baseurl_right = r"&limit=20"


savepath = ".\\豆瓣电影链接.xlsx"

Percentages = ['100%3A90','90%3A80','80%3A70','70%3A60','60%3A50','50%3A40','40%3A30','30%3A20','20%3A10','10%3A0']

def geturl():

    # 先爬剧情
    types = [str(24),str(5),str(13)]

    for type in types:
        print(type)
        for Percentage in Percentages:
            print("\t",Percentage)
            for i in range(0, 50):  # 调用获取页面信息的函数，50次
                print("\t\t",i)
                url_movie = []

                # 构造url
                url = baseurl_left + type + baseurl_midd1 + Percentage + baseurl_midd3 + str(i * 20) + baseurl_right
                html = askURL(url)
                # 生成随机休眠时间
                sleep_time = random.uniform(1, 3)  # 随机生成1到3秒的休眠时间
                time.sleep(sleep_time)  # 休眠一段时间

                 # 解析数据得到url
                try:
                    data = json.loads(html) #正确解析url,转化成列表，每个元素是一个字典，对应一部电影
                    print(data)
                    num = len(data)
                    print(num)
                    for j in range(0,num):
                        url_value = data[j].get("url")
                        url_movie.append(url_value)
                        if url_value:
                            print("url value:", url_value)
                        else:
                            print("url not found.")
                except json.JSONDecodeError as e:
                    print("Invalid JSON format:", e)
                except IndexError:
                    print("Index out of range error occurred while accessing list elements.")
                print(url_movie)


                # 每隔20次保存一次数据
                append_to_excel(url_movie, savepath)

    return url_movie


# 保存数据，追加
def append_to_excel(data_list, save_path):
    try:
        workbook = load_workbook(save_path)  # 加载现有的 Excel 文件
        sheet = workbook.active  # 获取活动工作表
        last_row = sheet.max_row  # 获取最后一行的行号
        start_row = last_row + 1  # 下一行开始写入的行号

        for data in data_list:
            sheet.append([data])  # 将数据封装到列表中并追加到工作表的下一行

        workbook.save(save_path)  # 保存更改
        workbook.close()  # 关闭工作簿
        print("Data appended to", save_path)
    except FileNotFoundError:
        print("Excel file not found. Creating a new file...")
        workbook = Workbook()  # 创建一个新的工作簿
        sheet = workbook.active  # 获取活动工作表
        sheet.insert_cols(1)  # 在第一列之前插入新列
        sheet.cell(row=1, column=1, value="电影链接URL")  # 设置第一行第一列的单元格值
        for data in data_list:
            sheet.append([data])  # 将数据封装到列表中并追加到工作表的下一行
        workbook.save(save_path)  # 保存工作簿
        workbook.close()  # 关闭工作簿
        print("New Excel file created and data appended to", save_path)
    except Exception as e:
        print("An error occurred while appending data:", e)



#得到指定url的网页内容
def askURL(url):
    head = {    #模拟浏览器信息，向豆瓣服务器发送消息
        "User-Agent":"Mozilla / 5.0(Windows NT 10.0;Win64;x64) AppleWebKit / 537.36(KHTML, likeGecko) Chrome / 97.0.4692.99 Safari / 537.36 Edg / 97.0.1072.76",
        "Referer":"https: // movie.douban.com / explore",
        "Cookie":'''bid=LX8SNrZjUco; __utmc=30149280; __utmz=30149280.1684685436.1.1.utmcsr=link.csdn.net|utmccn=(referral)|utmcmd=referral|utmcct = /; ll = "118318";__gads = ID = 58d9db24248a4833 - 228b36c62ce100d4: T = 1684686381:RT = 1684686381:S = ALNI_MaOE3fHztDZdzEeREgvL0Al - -nJSg;ct = y;ap_v = 0, 6.0;__utma = 30149280.1464348820.1684685436.1684685436.1684716053.2;__gpi = UID = 00000c094a8e76ab: T = 1684686381:RT = 1684716054:S = ALNI_MZP9Cn0IKLAC7WjqoSRIgsRr16v4g;__utmb = 30149280.1.10.1684716053'''
    }

    #用户代理，表示告诉服务器我们的机器类型（本质上告诉浏览器我们可以接受什么水平的内容）

    #发送消息
    request = urllib.request.Request(url=url, headers=head)  #Request封装了一下请求
    html = ""
    try:
        response = urllib.request.urlopen(request)  #接收返回的信息

        html = response.read().decode("utf-8")  #读取信息
        print(html)
    except urllib.error.URLError as e:
        if hasattr(e,"code"):       #该函数用于判断是否含指定字符串
            print(e.code)       #用上面的函数打印指定字符串
        if hasattr(e,"reason"):
            print(e.reason)
    return html


def main():
    geturl()


if __name__ == '__main__':
    main()